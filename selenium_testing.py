import openpyxl
from selenium import webdriver
import time

driver = webdriver.Chrome('E:\Selenium\chromedriver.exe')



# driver.find_element_by_xpath('/html/body/input').click()

# if driver.current_url == 'http://127.0.0.1:5000/':
#     print("\nTest 4 Passed: Page Loaded Successfully: ",driver.current_url)
# else:
#     print("\nTest 4 Failed: Wrong Page Loaded: ",driver.current_url)

path = r"D:\\FINAL MINI PROJECT\\testcases.xlsx"

def writedata(file, sheetName, rownum, columnno, data):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	sheet.cell(row=rownum, column = columnno).value = data
	workbook.save(file)

# wk = openpyxl.load_workbook("D:\\FINAL MINI PROJECT\\testcases.xlsx")
# print(wk.sheetnames)
# print("Active sheet is:"+ wk.active.title)
# sh=wk['Sheet1']
# print(sh.title)

#Test case 1
driver.get(' http://127.0.0.1:5000/')
driver.maximize_window()

if driver.title == "Main Page":
	print("Main page loaded succesfully: test passed")
	writedata(path,"Sheet1", 2, 2, "test passed")
else:
	print("Main page did not load: test failed")
	writedata(path,"Sheet1", 2, 2, "test failed")
time.sleep(3)

#test case 2
driver.find_element_by_xpath('/html/body/div/div/form/input[1]').send_keys(r'D:\FINAL MINI PROJECT\data.txt')
existance =driver.find_element_by_name('file')
if(existance):
	print("Data set loaded succesfully: test passed")
	writedata(path,"Sheet1", 3, 2, "test passed")
else:
	print("Dataset did not load: test failed")
	writedata(path,"Sheet1", 3, 2, "test failed")



#Test case 3
driver.find_element_by_xpath('/html/body/div/div/form/input[2]').click()
if driver.title == "Cluster Table":
	print("Cluster table displayed succesfully: test passed")
	writedata(path,"Sheet1", 4, 2, "test passed")
else:
	print("Cluster table not displayed: test failed")
	writedata(path,"Sheet1", 4, 2, "test failed")
time.sleep(3)

#Test case 4
driver.find_element_by_name('show_clusters').click()
if driver.title == "Clusters":
	print("Clusters plotted succesfully: test passed")
	writedata(path,"Sheet1", 5, 2, "test passed")
else:
	print("Clusters were not displayed: test failed")
	writedata(path,"Sheet1", 5, 2, "test failed")
time.sleep(3)


#Test case 5
driver.find_element_by_name('Home').click()
if driver.title == "Main Page":
	print("Redirected to Home page: test passed")
	writedata(path,"Sheet1", 6, 2, "test passed")
else:
	print("Redirection unsuccesful: test failed")
	writedata(path,"Sheet1", 6, 2, "test failed")
#     print("\nTest 2 Passed:\nElement exists:",existance.text,"\nDisplayed:",existance.is_displayed(),"\nEnabled:",existance.is_enabled())
# else:
#     print("\nTest 2 Failed:\nElement doesn't exist")



# filename = driver.find_element_by_name("file")
# if filename == 'data.txt':
# 	print("test passed")
# 	writedata(path,"Sheet1", 3, 2, "test passed")
# else:
# 	print("test failed")
# 	writedata(path,"Sheet1", 3, 2, "test failed")	